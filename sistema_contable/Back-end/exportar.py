# ============================================================
# exportar.py
# Sistema Contable - Exportación a Excel y PDF
# Corregido: solo cuentas de 2 dígitos
# ============================================================

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from libro_diario import obtener_libro_diario
from libro_mayor import obtener_mayor_agrupado
from balance_comprobacion import obtener_balance_comprobacion
from estado_situacion_financiera import obtener_estado_situacion_financiera
from estado_resultados import obtener_estado_resultados

COLOR_HEADER  = "1F4E79"
COLOR_SUBHEAD = "2E75B6"
COLOR_ALT_ROW = "DEEAF1"
COLOR_TOTAL   = "FFD700"
COLOR_WHITE   = "FFFFFF"
COLOR_SEC     = "366092"

_thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

_AZUL    = colors.HexColor("#1F4E79")
_AZUL_M  = colors.HexColor("#2E75B6")
_AZUL_CL = colors.HexColor("#DEEAF1")
_AZUL_S  = colors.HexColor("#366092")
_DORADO  = colors.HexColor("#FFD700")
_GRIS    = colors.HexColor("#F5F5F5")

styles       = getSampleStyleSheet()
_title_style = ParagraphStyle(
    "titulo", parent=styles["Normal"],
    fontName="Helvetica-Bold", fontSize=13,
    textColor=colors.white, backColor=_AZUL,
    alignment=TA_CENTER, spaceAfter=8, spaceBefore=4,
    leftIndent=8, rightIndent=8
)
_section_style = ParagraphStyle(
    "seccion", parent=styles["Normal"],
    fontName="Helvetica-Bold", fontSize=10,
    textColor=colors.white, backColor=_AZUL_S,
    alignment=TA_LEFT, spaceAfter=2, spaceBefore=4,
    leftIndent=6
)


# ════════════════════════════════════════════════════════════
# HELPERS EXCEL
# ════════════════════════════════════════════════════════════

def _eh(ws, fila, cols, titulo, color=COLOR_HEADER):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=cols)
    c = ws.cell(row=fila, column=1, value=titulo)
    c.font      = Font(bold=True, color=COLOR_WHITE, size=12)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _ec(ws, fila, encabs, color=COLOR_SUBHEAD):
    for col, t in enumerate(encabs, 1):
        c = ws.cell(row=fila, column=col, value=t)
        c.font      = Font(bold=True, color=COLOR_WHITE)
        c.fill      = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center")
        c.border    = _thin_border


def _fd(ws, fila, datos, alt=False):
    bg = COLOR_ALT_ROW if alt else COLOR_WHITE
    for col, v in enumerate(datos, 1):
        c = ws.cell(row=fila, column=col, value=v)
        c.fill   = PatternFill("solid", fgColor=bg)
        c.border = _thin_border
        if isinstance(v, float):
            c.number_format = '#,##0.00'
            c.alignment     = Alignment(horizontal="right")


def _ft(ws, fila, datos):
    for col, v in enumerate(datos, 1):
        c = ws.cell(row=fila, column=col, value=v)
        c.font   = Font(bold=True)
        c.fill   = PatternFill("solid", fgColor=COLOR_TOTAL)
        c.border = _thin_border
        if isinstance(v, float):
            c.number_format = '#,##0.00'
            c.alignment     = Alignment(horizontal="right")


def _aa(ws, n):
    for col in range(1, n + 1):
        mx = 0
        cl = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if cell.value:
                        mx = max(mx, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[cl].width = min(mx + 4, 45)


# ════════════════════════════════════════════════════════════
# HOJAS EXCEL
# ════════════════════════════════════════════════════════════

def _hoja_libro_diario(wb):
    ws  = wb.create_sheet("Libro Diario")
    enc = ["Asiento", "Fecha", "Glosa", "Cta.", "Nombre Cuenta", "D/H", "DEBE (S/)", "HABER (S/)"]
    _eh(ws, 1, len(enc), "LIBRO DIARIO")
    _ec(ws, 2, enc)
    fila = 3; alt = False; td = th = 0.0
    for d in obtener_libro_diario():
        _fd(ws, fila, [
            d["cod_asiento"], d["fecha"], d["glosa"],
            d["cod_cuenta"],  d["nombre_cuenta"], d["tipo_movimiento"],
            d["debe"] or None, d["haber"] or None
        ], alt)
        td += d["debe"]; th += d["haber"]; alt = not alt; fila += 1
    _ft(ws, fila, ["", "", "", "", "", "TOTAL", round(td, 2), round(th, 2)])
    _aa(ws, len(enc))


def _hoja_libro_mayor(wb):
    ws   = wb.create_sheet("Libro Mayor")
    fila = 1
    for cuenta in obtener_mayor_agrupado():
        enc = ["Fecha", "Glosa", "DEBE (S/)", "HABER (S/)", "Saldo Acum."]
        _eh(ws, fila, len(enc), f"[{cuenta['cod_cuenta']}] {cuenta['nombre_cuenta']}")
        fila += 1; _ec(ws, fila, enc); fila += 1; alt = False
        for mov in cuenta["movimientos"]:
            _fd(ws, fila, [
                mov["fecha"], mov["glosa"],
                mov["debe"]  or None,
                mov["haber"] or None,
                mov["saldo_acumulado"]
            ], alt)
            alt = not alt; fila += 1
        _ft(ws, fila, [
            "TOTAL", "",
            cuenta["total_debe"], cuenta["total_haber"], cuenta["saldo_final"]
        ])
        fila += 2
    _aa(ws, 5)


def _hoja_balance(wb):
    ws  = wb.create_sheet("Balance Comprobación")
    bal = obtener_balance_comprobacion()
    enc = ["Cta.", "Nombre de la Cuenta", "DEBE (S/)", "HABER (S/)"]
    _eh(ws, 1, len(enc), f"BALANCE DE COMPROBACIÓN  ·  {bal['periodo']}")
    _ec(ws, 2, enc)
    fila = 3; alt = False
    for d in bal["detalle"]:
        _fd(ws, fila, [
            d["cod_cuenta"], d["nombre_cuenta"],
            d["debe"] or None, d["haber"] or None
        ], alt)
        alt = not alt; fila += 1
    _ft(ws, fila, ["", "TOTAL", bal["total_debe"], bal["total_haber"]])
    ws.cell(row=fila + 1, column=1, value=f"Estado: {bal['estado']}").font = Font(bold=True)
    _aa(ws, len(enc))


def _hoja_esf(wb):
    ws  = wb.create_sheet("Situación Financiera")
    esf = obtener_estado_situacion_financiera()

    for col, w in {"A": 6, "B": 30, "C": 15, "D": 2, "E": 30, "F": 15}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    c        = ws["A1"]
    c.value  = f"ESTADO DE SITUACIÓN FINANCIERA  ·  Al {esf['fecha_corte']}"
    c.font   = Font(bold=True, color=COLOR_WHITE, size=13)
    c.fill   = PatternFill("solid", fgColor=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    for col, txt in [(1, "ACTIVO"), (5, "PASIVO Y PATRIMONIO")]:
        c       = ws.cell(row=2, column=col, value=txt)
        c.font  = Font(bold=True, color=COLOR_WHITE)
        c.fill  = PatternFill("solid", fgColor=COLOR_SUBHEAD)
        c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)

    fi = fd = 3

    def _izq(titulo, items, lbl, total):
        nonlocal fi
        ws.merge_cells(start_row=fi, start_column=1, end_row=fi, end_column=3)
        c      = ws.cell(row=fi, column=1, value=titulo)
        c.font = Font(bold=True, color=COLOR_WHITE)
        c.fill = PatternFill("solid", fgColor=COLOR_SEC)
        fi    += 1
        alt    = False
        for item in items:
            ws.cell(row=fi, column=2, value=item["nombre_cuenta"])
            c               = ws.cell(row=fi, column=3, value=item["valor"])
            c.number_format = '#,##0.00'
            c.alignment     = Alignment(horizontal="right")
            if alt:
                for cl in [1, 2, 3]:
                    ws.cell(row=fi, column=cl).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
            alt = not alt; fi += 1
        ws.cell(row=fi, column=2, value=lbl).font = Font(bold=True)
        c               = ws.cell(row=fi, column=3, value=total)
        c.font          = Font(bold=True)
        c.number_format = '#,##0.00'
        c.alignment     = Alignment(horizontal="right")
        c.fill          = PatternFill("solid", fgColor="BDD7EE")
        fi += 1

    def _der(titulo, items, lbl, total):
        nonlocal fd
        ws.merge_cells(start_row=fd, start_column=5, end_row=fd, end_column=6)
        c      = ws.cell(row=fd, column=5, value=titulo)
        c.font = Font(bold=True, color=COLOR_WHITE)
        c.fill = PatternFill("solid", fgColor=COLOR_SEC)
        fd    += 1
        alt    = False
        for item in items:
            ws.cell(row=fd, column=5, value=item["nombre_cuenta"])
            c               = ws.cell(row=fd, column=6, value=item["valor"])
            c.number_format = '#,##0.00'
            c.alignment     = Alignment(horizontal="right")
            if alt:
                for cl in [5, 6]:
                    ws.cell(row=fd, column=cl).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
            alt = not alt; fd += 1
        ws.cell(row=fd, column=5, value=lbl).font = Font(bold=True)
        c               = ws.cell(row=fd, column=6, value=total)
        c.font          = Font(bold=True)
        c.number_format = '#,##0.00'
        c.alignment     = Alignment(horizontal="right")
        c.fill          = PatternFill("solid", fgColor="BDD7EE")
        fd += 1

    _izq("ACTIVO CORRIENTE",    esf["activo"]["corriente"],
         "Total Activo Corriente",    esf["activo"]["total_corriente"])
    _izq("ACTIVO NO CORRIENTE", esf["activo"]["no_corriente"],
         "Total Activo No Corriente", esf["activo"]["total_no_corriente"])
    _der("PASIVO CORRIENTE",    esf["pasivo"]["corriente"],
         "Total Pasivo Corriente",    esf["pasivo"]["total_corriente"])
    _der("PASIVO NO CORRIENTE", esf["pasivo"]["no_corriente"],
         "Total Pasivo No Corriente", esf["pasivo"]["total_no_corriente"])
    _der("PATRIMONIO",          esf["patrimonio"]["detalle"],
         "Total Patrimonio",          esf["patrimonio"]["total_patrimonio"])

    ft = max(fi, fd) + 1
    ws.cell(row=ft, column=2, value="TOTAL ACTIVO").font = Font(bold=True)
    c               = ws.cell(row=ft, column=3, value=esf["activo"]["total_activo"])
    c.font          = Font(bold=True)
    c.number_format = '#,##0.00'
    c.fill          = PatternFill("solid", fgColor=COLOR_TOTAL)
    c.alignment     = Alignment(horizontal="right")
    ws.cell(row=ft, column=5, value="TOTAL PASIVO + PATRIMONIO").font = Font(bold=True)
    c               = ws.cell(row=ft, column=6, value=esf["total_pasivo_patrimonio"])
    c.font          = Font(bold=True)
    c.number_format = '#,##0.00'
    c.fill          = PatternFill("solid", fgColor=COLOR_TOTAL)
    c.alignment     = Alignment(horizontal="right")


def _hoja_resultados(wb):
    ws  = wb.create_sheet("Estado de Resultados")
    er  = obtener_estado_resultados()
    enc = ["Descripción", "Monto (S/)"]
    _eh(ws, 1, len(enc),
        f"ESTADO DE RESULTADOS  ·  {er['periodo']['fecha_inicio']} al {er['periodo']['fecha_fin']}")
    _ec(ws, 2, enc)
    filas = [
        ("n", "Ventas Netas / Ingresos por Servicios",   er["ingresos"]["ventas_netas"]),
        ("n", "(-) Costo de Ventas",                     -er["costo_ventas"]["monto"]),
        ("t", "UTILIDAD BRUTA",                           er["utilidad_bruta"]),
        ("n", "(-) Gastos de Ventas",                    -er["gastos_operativos"]["gastos_ventas"]["monto"]),
        ("n", "(-) Gastos de Administración",             -er["gastos_operativos"]["gastos_administracion"]["monto"]),
        ("t", "UTILIDAD OPERATIVA",                       er["utilidad_operativa"]),
        ("n", "(+) Otros Ingresos",                       er["otros"]["otros_ingresos"]),
        ("n", "(-) Gastos Financieros",                  -er["otros"]["gastos_financieros"]),
        ("n", "(+) Ingresos Financieros",                 er["otros"]["ingresos_financieros"]),
        ("t", "UTILIDAD ANTES DE IMPUESTOS",              er["utilidad_antes_impuestos"]),
        ("n", f"(-) Impuesto a la Renta ({er['impuesto_renta']['tasa']*100:.1f}%)",
              -er["impuesto_renta"]["monto"]),
        ("t", f"{er['resultado']} NETA DEL PERÍODO",     er["utilidad_neta"]),
    ]
    totales = {2, 5, 8, 11}
    for idx, (tipo, desc, monto) in enumerate(filas):
        row = idx + 3
        if idx in totales:
            _ft(ws, row, [desc, monto])
        else:
            _fd(ws, row, [desc, monto], idx % 2 == 0)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18


# ════════════════════════════════════════════════════════════
# API EXCEL
# ════════════════════════════════════════════════════════════

def exportar_excel(tablas=None):
    if tablas is None:
        tablas = ["diario", "mayor", "balance", "situacion", "resultados"]
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    mapa = {
        "diario":     _hoja_libro_diario,
        "mayor":      _hoja_libro_mayor,
        "balance":    _hoja_balance,
        "situacion":  _hoja_esf,
        "resultados": _hoja_resultados,
    }
    for tabla in tablas:
        if tabla in mapa:
            mapa[tabla](wb)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════
# HELPERS PDF
# ════════════════════════════════════════════════════════════

def _tp(datos, encabs, col_widths=None):
    t = Table([encabs] + datos, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), _AZUL_M),
        ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
        ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, 0), 8),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",       (0, 1), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _AZUL_CL]),
    ]))
    return t


def _tot_row(filas, col_widths):
    t = Table([filas], colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _DORADO),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
    ]))
    return t


# ════════════════════════════════════════════════════════════
# SECCIONES PDF
# ════════════════════════════════════════════════════════════

def _pdf_diario():
    datos = obtener_libro_diario()
    el    = [Paragraph("LIBRO DIARIO", _title_style), Spacer(1, 0.2 * cm)]
    enc   = ["Asiento", "Fecha", "Glosa", "Cta.", "Cuenta", "D/H", "DEBE (S/)", "HABER (S/)"]
    cw    = [1.2*cm, 2*cm, 5.5*cm, 1*cm, 5*cm, 0.8*cm, 2.5*cm, 2.5*cm]
    filas = []; td = th = 0.0
    for d in datos:
        filas.append([
            str(d["cod_asiento"]), d["fecha"], d["glosa"][:35],
            str(d["cod_cuenta"]),  d["nombre_cuenta"][:28],
            d["tipo_movimiento"],
            f"{d['debe']:,.2f}"  if d["debe"]  else "",
            f"{d['haber']:,.2f}" if d["haber"] else ""
        ])
        td += d["debe"]; th += d["haber"]
    el.append(_tp(filas, enc, cw))
    el.append(_tot_row(["", "", "", "", "", "TOTAL", f"{td:,.2f}", f"{th:,.2f}"], cw))
    return el


def _pdf_mayor():
    mayor = obtener_mayor_agrupado()
    el    = [Paragraph("LIBRO MAYOR", _title_style), Spacer(1, 0.2 * cm)]
    cw    = [2*cm, 9*cm, 3*cm, 3*cm, 3*cm]
    for cuenta in mayor:
        el.append(Paragraph(
            f"[{cuenta['cod_cuenta']}] {cuenta['nombre_cuenta']}", _section_style
        ))
        enc   = ["Fecha", "Glosa", "DEBE (S/)", "HABER (S/)", "Saldo"]
        filas = [
            [
                m["fecha"], m["glosa"][:40],
                f"{m['debe']:,.2f}"  if m["debe"]  else "",
                f"{m['haber']:,.2f}" if m["haber"] else "",
                f"{m['saldo_acumulado']:,.2f}"
            ]
            for m in cuenta["movimientos"]
        ]
        el.append(_tp(filas, enc, cw))
        el.append(_tot_row([
            "TOTAL", "",
            f"{cuenta['total_debe']:,.2f}",
            f"{cuenta['total_haber']:,.2f}",
            f"{cuenta['saldo_final']:,.2f}"
        ], cw))
        el.append(Spacer(1, 0.3 * cm))
    return el


def _pdf_balance():
    bal = obtener_balance_comprobacion()
    el  = [
        Paragraph(f"BALANCE DE COMPROBACIÓN  ·  {bal['periodo']}", _title_style),
        Spacer(1, 0.2 * cm)
    ]
    cw    = [1.5*cm, 11*cm, 3.5*cm, 3.5*cm]
    filas = [
        [str(d["cod_cuenta"]), d["nombre_cuenta"],
         f"{d['debe']:,.2f}", f"{d['haber']:,.2f}"]
        for d in bal["detalle"]
    ]
    el.append(_tp(filas, ["Cta.", "Nombre de Cuenta", "DEBE (S/)", "HABER (S/)"], cw))
    el.append(_tot_row(
        ["", "TOTAL", f"{bal['total_debe']:,.2f}", f"{bal['total_haber']:,.2f}"], cw
    ))
    col = (colors.HexColor("#27AE60") if bal["estado"] == "CUADRADO"
           else colors.HexColor("#E74C3C"))
    st  = ParagraphStyle("e", fontSize=9, fontName="Helvetica-Bold",
                          textColor=col, spaceBefore=4)
    el.append(Paragraph(f"Estado: {bal['estado']}", st))
    return el


def _pdf_esf():
    esf = obtener_estado_situacion_financiera()
    el  = [
        Paragraph(
            f"ESTADO DE SITUACIÓN FINANCIERA  ·  Al {esf['fecha_corte']}",
            _title_style
        ),
        Spacer(1, 0.3 * cm)
    ]

    def _build(secciones):
        filas = []
        for titulo, items, lbl, total in secciones:
            filas.append(("T", titulo, ""))
            for item in items:
                filas.append(("I", item["nombre_cuenta"], f"S/ {item['valor']:,.2f}"))
            filas.append(("S", lbl, f"S/ {total:,.2f}"))
        return filas

    col_a = _build([
        ("ACTIVO CORRIENTE",    esf["activo"]["corriente"],
         "Total Activo Corriente",    esf["activo"]["total_corriente"]),
        ("ACTIVO NO CORRIENTE", esf["activo"]["no_corriente"],
         "Total Activo No Corriente", esf["activo"]["total_no_corriente"]),
    ])
    col_p = _build([
        ("PASIVO CORRIENTE",    esf["pasivo"]["corriente"],
         "Total Pasivo Corriente",    esf["pasivo"]["total_corriente"]),
        ("PASIVO NO CORRIENTE", esf["pasivo"]["no_corriente"],
         "Total Pasivo No Corriente", esf["pasivo"]["total_no_corriente"]),
        ("PATRIMONIO",          esf["patrimonio"]["detalle"],
         "Total Patrimonio",          esf["patrimonio"]["total_patrimonio"]),
    ])
    col_p.append((
        "I", "Resultado del Ejercicio",
        f"S/ {esf['patrimonio']['resultado_ejercicio']:,.2f}"
    ))

    ml = max(len(col_a), len(col_p))
    while len(col_a) < ml: col_a.append(("V", "", ""))
    while len(col_p) < ml: col_p.append(("V", "", ""))

    AW = 8.5*cm; MW = 3*cm; SW = 0.3*cm
    cw = [AW, MW, SW, AW, MW]

    filas_tabla = [["ACTIVO", "", "", "PASIVO Y PATRIMONIO", ""]]
    estilos = [
        ("BACKGROUND", (0, 0), (1, 0), _AZUL_M), ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
        ("BACKGROUND", (3, 0), (4, 0), _AZUL_M), ("TEXTCOLOR", (3, 0), (4, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("FONTSIZE",   (0, 1), (-1, -1), 8), ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ALIGN",      (1, 0), (1, -1), "RIGHT"), ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("GRID",       (3, 0), (4, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("BACKGROUND", (2, 0), (2, -1), colors.white),
    ]

    for i, (fa, fp) in enumerate(zip(col_a, col_p), 1):
        filas_tabla.append([fa[1], fa[2], "", fp[1], fp[2]])
        if fa[0] == "T":
            estilos += [("BACKGROUND", (0, i), (1, i), _AZUL_S),
                        ("TEXTCOLOR",  (0, i), (1, i), colors.white),
                        ("FONTNAME",   (0, i), (1, i), "Helvetica-Bold")]
        elif fa[0] == "S":
            estilos += [("BACKGROUND", (0, i), (1, i), _AZUL_CL),
                        ("FONTNAME",   (0, i), (1, i), "Helvetica-Bold")]
        elif fa[0] == "I" and i % 2 == 0:
            estilos.append(("BACKGROUND", (0, i), (1, i), _GRIS))

        if fp[0] == "T":
            estilos += [("BACKGROUND", (3, i), (4, i), _AZUL_S),
                        ("TEXTCOLOR",  (3, i), (4, i), colors.white),
                        ("FONTNAME",   (3, i), (4, i), "Helvetica-Bold")]
        elif fp[0] == "S":
            estilos += [("BACKGROUND", (3, i), (4, i), _AZUL_CL),
                        ("FONTNAME",   (3, i), (4, i), "Helvetica-Bold")]
        elif fp[0] == "I" and i % 2 == 0:
            estilos.append(("BACKGROUND", (3, i), (4, i), _GRIS))

    fi_tot = len(filas_tabla)
    filas_tabla.append([
        "TOTAL ACTIVO", f"S/ {esf['activo']['total_activo']:,.2f}",
        "", "TOTAL PASIVO + PATRIMONIO",
        f"S/ {esf['total_pasivo_patrimonio']:,.2f}"
    ])
    estilos += [
        ("BACKGROUND", (0, fi_tot), (1, fi_tot), _DORADO),
        ("BACKGROUND", (3, fi_tot), (4, fi_tot), _DORADO),
        ("FONTNAME",   (0, fi_tot), (-1, fi_tot), "Helvetica-Bold")
    ]

    t = Table(filas_tabla, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle(estilos))
    el.append(t)

    cuadrado = esf.get("cuadrado", False)
    col      = colors.HexColor("#27AE60") if cuadrado else colors.HexColor("#E74C3C")
    msg      = ("✓ El Estado de Situación Financiera CUADRA" if cuadrado
                else f"✗ No cuadra — Diferencia: S/ {esf.get('diferencia', 0):,.2f}")
    el.append(Paragraph(msg, ParagraphStyle(
        "e2", fontSize=9, fontName="Helvetica-Bold",
        textColor=col, spaceBefore=6
    )))
    return el


def _pdf_resultados():
    er = obtener_estado_resultados()
    el = [
        Paragraph(
            f"ESTADO DE RESULTADOS  ·  {er['periodo']['fecha_inicio']} al {er['periodo']['fecha_fin']}",
            _title_style
        ),
        Spacer(1, 0.2 * cm)
    ]
    cw        = [13*cm, 5*cm]
    filas_raw = [
        ("n", "Ventas Netas / Ingresos por Servicios",  er["ingresos"]["ventas_netas"]),
        ("n", "(-) Costo de Ventas",                    er["costo_ventas"]["monto"]),
        ("t", "UTILIDAD BRUTA",                          er["utilidad_bruta"]),
        ("n", "(-) Gastos de Ventas",                   er["gastos_operativos"]["gastos_ventas"]["monto"]),
        ("n", "(-) Gastos de Administración",            er["gastos_operativos"]["gastos_administracion"]["monto"]),
        ("t", "UTILIDAD OPERATIVA",                      er["utilidad_operativa"]),
        ("n", "(+) Otros Ingresos",                      er["otros"]["otros_ingresos"]),
        ("n", "(-) Gastos Financieros",                  er["otros"]["gastos_financieros"]),
        ("n", "(+) Ingresos Financieros",                er["otros"]["ingresos_financieros"]),
        ("t", "UTILIDAD ANTES DE IMPUESTOS",             er["utilidad_antes_impuestos"]),
        ("n", f"(-) Impuesto a la Renta ({er['impuesto_renta']['tasa']*100:.1f}%)",
              er["impuesto_renta"]["monto"]),
        ("g", f"{er['resultado']} NETA DEL PERÍODO",    er["utilidad_neta"]),
    ]
    filas   = [[f[1], f"S/ {f[2]:,.2f}"] for f in filas_raw]
    t       = Table([["Descripción", "Monto (S/)"]] + filas, colWidths=cw, repeatRows=1)
    estilos = [
        ("BACKGROUND", (0, 0), (-1, 0), _AZUL_M), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN",      (1, 0), (1, -1), "RIGHT"), ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("FONTSIZE",   (0, 1), (-1, -1), 8), ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _AZUL_CL]),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
    ]
    for i, (tipo, _, __) in enumerate(filas_raw, 1):
        if tipo == "t":
            estilos += [("BACKGROUND", (0, i), (-1, i), _AZUL_CL),
                        ("FONTNAME",   (0, i), (-1, i), "Helvetica-Bold")]
        elif tipo == "g":
            estilos += [("BACKGROUND", (0, i), (-1, i), _DORADO),
                        ("FONTNAME",   (0, i), (-1, i), "Helvetica-Bold"),
                        ("FONTSIZE",   (0, i), (-1, i), 9)]
    t.setStyle(TableStyle(estilos))
    el.append(t)
    return el


# ════════════════════════════════════════════════════════════
# API PDF
# ════════════════════════════════════════════════════════════

def exportar_pdf(tablas=None):
    if tablas is None:
        tablas = ["diario", "mayor", "balance", "situacion", "resultados"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm,   bottomMargin=1.5*cm
    )
    mapa = {
        "diario":     _pdf_diario,
        "mayor":      _pdf_mayor,
        "balance":    _pdf_balance,
        "situacion":  _pdf_esf,
        "resultados": _pdf_resultados,
    }
    story = []
    for i, tabla in enumerate(tablas):
        if tabla in mapa:
            story.extend(mapa[tabla]())
            if i < len(tablas) - 1:
                story.append(PageBreak())
    doc.build(story)
    buf.seek(0)
    return buf.read()